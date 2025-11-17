# -*- coding: utf-8 -*-
"""
Rellena la hoja "Identificación del Instrumento" de una plantilla Excel
usando un archivo de datos (raw.xlsx/.csv) generado por 307-extract_camiones.py
y un YAML de mapeo campo→celda. Todo se hace con diálogos de Windows.

Requisitos:
    pip install pandas openpyxl pyyaml

Uso:
    - Doble clic en Windows, o
    - python fill_identificacion_gui.py
"""

from pathlib import Path
import sys
import traceback
import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter, column_index_from_string


# --- GUI (tkinter) ---
import tkinter as tk
from tkinter import filedialog, messagebox

try:
    import yaml
except Exception:
    tk.Tk().withdraw()
    messagebox.showerror("Falta dependencia", "No se encontró PyYAML.\nInstalá con: pip install pyyaml")
    sys.exit(1)

DEFAULT_SHEET = "Identificación del Instrumento"


def load_map(yaml_path: Path) -> dict:
    with open(yaml_path, "r", encoding="utf-8") as f:
        mapping = yaml.safe_load(f) or {}
    norm = {}
    for k, v in mapping.items():
        if isinstance(v, str):
            norm[k] = {"cell": v, "keep_existing": False}
        elif isinstance(v, dict) and "cell" in v:
            norm[k] = {"cell": str(v["cell"]), "keep_existing": bool(v.get("keep_existing", False))}
        else:
            raise ValueError(f"Entrada de mapeo inválida para '{k}': {v}")
    return norm

def _anchor_of_merged_cell(ws, cell_ref: str) -> str:
    """
    Si cell_ref pertenece a un rango combinado, devuelve la referencia
    de su ancla (esquina superior izquierda). Si no, devuelve cell_ref.
    """
    col_letters = ''.join(c for c in cell_ref if c.isalpha())
    row_numbers = ''.join(c for c in cell_ref if c.isdigit())
    col_idx = column_index_from_string(col_letters)
    row_idx = int(row_numbers)

    for merged in ws.merged_cells.ranges:
        if (merged.min_row <= row_idx <= merged.max_row) and (merged.min_col <= col_idx <= merged.max_col):
            return f"{get_column_letter(merged.min_col)}{merged.min_row}"
    return cell_ref


def fill_sheet(template_path: Path, out_path: Path, mapping: dict, record: dict, sheet_name: str = DEFAULT_SHEET):
    wb = load_workbook(template_path)
    if sheet_name not in wb.sheetnames:
        raise KeyError(f"La hoja '{sheet_name}' no existe en la plantilla. Hojas: {wb.sheetnames}")
    ws = wb[sheet_name]

    for field, cfg in mapping.items():
        cell = cfg["cell"]
        keep_existing = cfg["keep_existing"]
        value = record.get(field, "")

        # Saltar valores vacíos → conserva lo que ya tenga la plantilla (fechas, etc.)
        if value is None or str(value).strip() == "":
            continue

        # Si la celda está en un rango combinado, escribir en el ANCLA (arriba-izquierda)
        anchor = _anchor_of_merged_cell(ws, cell)

        # Respetar keep_existing mirando el valor del ancla
        existing = ws[anchor].value
        if keep_existing and (existing is not None) and (str(existing).strip() != ""):
            continue

        ws[anchor] = value

    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_path)

def suggest_out_name(df: pd.DataFrame) -> str:
    try:
        if "Numero de O.T." in df.columns:
            ot_val = str(df["Numero de O.T."].iloc[0]).strip()
            if ot_val:
                return f"OT_{ot_val}_final.xlsx"
    except Exception:
        pass
    return "OT_final.xlsx"


def main():
    root = tk.Tk()
    root.withdraw()  # ventana invisible, solo usamos diálogos

    try:
        # 1) Seleccionar ORIGEN (raw .xlsx/.csv)
        raw_path_str = filedialog.askopenfilename(
            title="Seleccioná el Excel/CSV ORIGEN (datos del scraper)",
            filetypes=[("Excel / CSV", "*.xlsx *.xlsm *.csv"), ("Todos", "*.*")]
        )
        if not raw_path_str:
            return
        raw_path = Path(raw_path_str)
        if raw_path.suffix.lower() in (".xlsx", ".xlsm"):
            df = pd.read_excel(raw_path)
        elif raw_path.suffix.lower() == ".csv":
            df = pd.read_csv(raw_path)
        else:
            messagebox.showerror("Archivo no válido", "El ORIGEN debe ser .xlsx/.xlsm o .csv")
            return
        if df.empty:
            messagebox.showinfo("Sin datos", "El archivo ORIGEN no tiene filas.")
            return

        # 2) Seleccionar PLANTILLA
        template_path_str = filedialog.askopenfilename(
            title="Seleccioná el Excel PLANTILLA a rellenar (.xlsx)",
            filetypes=[("Excel", "*.xlsx *.xlsm"), ("Todos", "*.*")]
        )
        if not template_path_str:
            return
        template_path = Path(template_path_str)

        # 3) Seleccionar MAPEO YAML
        map_path_str = filedialog.askopenfilename(
            title="Seleccioná el YAML de MAPEO campo→celda",
            filetypes=[("YAML", "*.yml *.yaml"), ("Todos", "*.*")]
        )
        if not map_path_str:
            return
        map_path = Path(map_path_str)
        mapping = load_map(map_path)

        # 4) ¿Multiples instrumentos?
        multi = False
        if len(df) > 1:
            multi = messagebox.askyesno(
                "Múltiples instrumentos",
                f"Se detectaron {len(df)} instrumentos.\n\n¿Querés generar un archivo por instrumento?"
            )

        # 5) Elegir SALIDA
        sugerido = suggest_out_name(df)
        out_path_str = filedialog.asksaveasfilename(
            title="Elegí dónde guardar el Excel RESULTADO",
            defaultextension=".xlsx",
            initialfile=sugerido,
            filetypes=[("Excel", "*.xlsx *.xlsm")]
        )
        if not out_path_str:
            return
        out_path = Path(out_path_str)

        # Confirma sobrescritura si existe (asksaveasfilename ya lo hace en la mayoría de shells,
        # pero lo reforzamos por si acaso)
        if out_path.exists():
            ok = messagebox.askyesno("Sobrescribir",
                                     f"El archivo '{out_path.name}' ya existe.\n\n¿Querés sobrescribirlo?")
            if not ok:
                # Permitir volver a elegir
                out_path_str = filedialog.asksaveasfilename(
                    title="Elegí otro nombre/ubicación para el Excel RESULTADO",
                    defaultextension=".xlsx",
                    initialfile=out_path.stem + " (1).xlsx",
                    filetypes=[("Excel", "*.xlsx *.xlsm")]
                )
                if not out_path_str:
                    return
                out_path = Path(out_path_str)

        # 6) Ejecutar llenado
        if multi:
            # genera múltiples archivos, uno por fila
            base = out_path.with_suffix("")  # quitar .xlsx
            count = 0
            for idx, row in df.iterrows():
                sufijo = f"_{idx+1}"
                serie = str(row.get("N° de serie Receptor", "")).strip()
                if serie and serie.lower() != "nan":
                    sufijo = f"_{serie}"
                destino = Path(str(base) + sufijo + out_path.suffix)
                try:
                    fill_sheet(template_path, destino, mapping, row.to_dict(), sheet_name=DEFAULT_SHEET)
                    count += 1
                except Exception as e:
                    messagebox.showwarning(
                        "Aviso",
                        f"No se pudo generar '{destino.name}'.\n\nDetalle:\n{e}"
                    )
            messagebox.showinfo("Listo", f"Generados {count} archivo(s) en:\n{out_path.parent}")
        else:
            # un único archivo con la primera fila (o elegís cuál; aquí usamos la primera)
            record = df.iloc[0].to_dict()
            fill_sheet(template_path, out_path, mapping, record, sheet_name=DEFAULT_SHEET)
            messagebox.showinfo("Listo", f"Generado:\n{out_path}")

    except Exception as e:
        traceback.print_exc()
        messagebox.showerror("Error inesperado", f"Ocurrió un error:\n\n{e}")


if __name__ == "__main__":
    main()
