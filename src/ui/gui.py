# -*- coding: utf-8 -*-
"""
GUI de extracción MetroWeb → Excel.

Ajustes visuales solicitados:
- Encabezado: "Extractor de datos (Beta)" con versión visible.
- Se elimina el badge "IDLE".
- Área de registro (log) con altura fija para evitar la "franja negra".
- Progreso gradual igual que antes.

Ejecución recomendada:
  python -m src.ui.gui
"""

# --- bootstrap robusto del proyecto (permite ejecutar este archivo "a pelo") ---
import os
import platform
import re
import sys
import threading
import subprocess
import xml.etree.ElementTree as ET
import zipfile
from datetime import datetime
from pathlib import Path
from typing import Any, Callable, Mapping, Optional, cast

import tkinter as tk
from tkinter import filedialog, messagebox, scrolledtext, ttk
from openpyxl import load_workbook


def find_project_root(markers=("pyproject.toml", "requirements.txt", ".git")) -> Path:
    """Busca la raiz del proyecto usando marcadores conocidos."""
    p = Path(__file__).resolve()
    for parent in (p.parent, *p.parents):
        if any((parent / m).exists() for m in markers):
            return parent
    # fallback por si no encuentra marcadores
    return Path(__file__).resolve().parents[2]


ROOT = find_project_root()
if str(ROOT) not in sys.path:
    sys.path.insert(0, str(ROOT))
from src.version import APP_VERSION

# -------------------------------------------------------------------------------

# ================== Config de app/estilos ==================

APP_NAME = "INTI METROWEB"
APP_SUBTITLE = "Extractor de datos (Beta)"
# Versión tomada automáticamente desde pyproject.toml (src/version.py)
# Ejemplo: APP_VERSION = "v0.4.0"
TEMPLATE_CAMION_PATH = ROOT / "assets" / "plantilla_camion.xlsx"


# Paleta
BG = "#f5f6f8"
CARD = "#ffffff"
PRIMARY = "#4472C4"
PRIMARY_DARK = "#365E9D"
MUTED = "#666"
SEPARATOR = "#e8e8e8"
SUCCESS = "#28a745"
DISABLED = "#BDBDBD"

# ================== Dependencias internas ==================

_MISSING_DEPS = []

try:
    import pandas as pd  # noqa: F401
except Exception as _e:  # pragma: no cover
    _MISSING_DEPS.append(f"pandas ({_e})")

try:
    from src.portal.scraper import extraer_camiones_por_ot
except Exception as _e:  # pragma: no cover
    _MISSING_DEPS.append(
        "src.portal.scraper.extraer_camiones_por_ot (verifica el archivo scraper.py y la firma)"
    )

try:
    from src.io.excel_exporter import (
        DATE_FIELDS,
        armar_hoja_verificacion_2columnas,
        _fecha_castellano,
    )
except Exception as _e:  # pragma: no cover
    _MISSING_DEPS.append(
        "src.io.excel_exporter (faltan armar_hoja_verificacion_2columnas y dependencias asociadas)"
    )

try:
    from src.domain.address import parse_domicilio_fiscal  # noqa: F401
except Exception as _e:
    _MISSING_DEPS.append(
        "src.domain.address.parse_domicilio_fiscal (revisa address.py)"
    )

# Import “a prueba de balas” del helper para anexar hoja "datos vpe"
append_sheet_as_first = None
try:
    from src.ui.excel_merge import (
        append_sheet_as_first,
    )  # ejecución como paquete (python -m src.ui.gui)
except Exception:
    try:
        from .excel_merge import (
            append_sheet_as_first,
        )  # ejecución directa desde src/ui (python gui.py)
    except Exception:
        append_sheet_as_first = None


# ================== Helpers Excel ==================


def _mapear_hojas(xlsx_path: Path) -> dict[str, str]:
    """Devuelve un mapa {nombre_hoja: target_xml} usando workbook.xml."""

    with zipfile.ZipFile(xlsx_path) as zf:
        wb_root = ET.fromstring(zf.read("xl/workbook.xml"))
        rel_root = ET.fromstring(zf.read("xl/_rels/workbook.xml.rels"))

    ns = {
        "ns": "http://schemas.openxmlformats.org/spreadsheetml/2006/main",
        "r": "http://schemas.openxmlformats.org/officeDocument/2006/relationships",
    }
    rels = {
        rel.attrib["Id"]: rel.attrib["Target"]
        for rel in rel_root.findall(
            ".//{http://schemas.openxmlformats.org/officeDocument/2006/relationships}Relationship"
        )
    }

    mapping: dict[str, str] = {}
    for sheet in wb_root.findall("ns:sheets/ns:sheet", ns):
        rel_id = sheet.attrib.get(
            "{http://schemas.openxmlformats.org/officeDocument/2006/relationships}id"
        )
        target = rels.get(rel_id)
        if target:
            mapping[sheet.attrib.get("name", "")] = target
    return mapping


def _restaurar_hojas_desde_template(
    *, template_path: Path, destino: Path, hojas_a_preservar: list[str]
) -> None:
    """
    Reemplaza la(s) hoja(s) solicitada(s) en ``destino`` con el XML de ``template_path``.

    Esto evita que se pierdan elementos no soportados por ``openpyxl`` (encabezado/pie
    con imágenes) tras guardar el archivo.
    """

    if not hojas_a_preservar:
        return

    mapping = _mapear_hojas(template_path)
    archivos_a_reemplazar = set()
    for hoja in hojas_a_preservar:
        target = mapping.get(hoja)
        if not target:
            continue
        hoja_path = f"xl/{target}"
        rel_path = f"xl/worksheets/_rels/{Path(target).name}.rels"
        archivos_a_reemplazar.add(hoja_path)
        archivos_a_reemplazar.add(rel_path)

    if not archivos_a_reemplazar:
        return

    tmp_path = destino.with_suffix(".tmp.xlsx")
    with (
        zipfile.ZipFile(destino, "r") as z_out,
        zipfile.ZipFile(template_path, "r") as z_tpl,
        zipfile.ZipFile(tmp_path, "w") as z_new,
    ):
        for nombre in z_out.namelist():
            if nombre in archivos_a_reemplazar and nombre in z_tpl.namelist():
                data = z_tpl.read(nombre)
            else:
                data = z_out.read(nombre)
            z_new.writestr(nombre, data)

    tmp_path.replace(destino)


# ================== Utilidades ==================


def limpiar_nombre_archivo(texto: str) -> str:
    """Normaliza un nombre de archivo reemplazando caracteres invalidos."""
    invalidos = ["<", ">", ":", '"', "/", "\\", "|", "?", "*"]
    for ch in invalidos:
        texto = texto.replace(ch, "_")
    return texto.strip()[:100] or "SIN_NOMBRE"


def validar_formato_ot(ot: str) -> bool:
    """Valida que la OT siga el patron NNN-NNNNN."""
    return bool(re.match(r"^\d{3}-\d{5}$", ot))


def validar_fecha_ddmmaaaa(fecha: str) -> bool:
    """Valida fechas en formato dd/mm/aaaa (ej. 16/12/2025)."""
    if not re.match(r"^\d{2}/\d{2}/\d{4}$", fecha):
        return False
    try:
        datetime.strptime(fecha, "%d/%m/%Y")
    except ValueError:
        return False
    return True


class ModernButton(tk.Canvas):
    def __init__(
        self,
        parent: tk.Widget,
        text: str,
        command: Optional[Callable[[], None]],
        bg_color: str = SUCCESS,
        fg_color: str = "white",
        hover_color: str = "#218838",
        width: int = 240,
        height: int = 44,
    ) -> None:
        super().__init__(
            parent, width=width, height=height, highlightthickness=0, bg=parent["bg"]
        )
        self._cmd = command
        self._bg = bg_color
        self._fg = fg_color
        self._hover = hover_color
        self.rect = self.create_rectangle(
            2, 2, width - 2, height - 2, fill=bg_color, width=0
        )
        self.text = self.create_text(
            width // 2,
            height // 2,
            text=text,
            fill=fg_color,
            font=("Segoe UI", 10, "bold"),
        )
        self.bind("<Enter>", lambda *_: self.itemconfig(self.rect, fill=self._hover))
        self.bind("<Leave>", lambda *_: self.itemconfig(self.rect, fill=self._bg))
        self.bind("<Button-1>", lambda *_: self._cmd and self._cmd())

    def set_enabled(self, enabled: bool) -> None:
        if enabled:
            self.itemconfig(self.rect, fill=self._bg)
            self.bind(
                "<Enter>", lambda *_: self.itemconfig(self.rect, fill=self._hover)
            )
            self.bind("<Leave>", lambda *_: self.itemconfig(self.rect, fill=self._bg))
            self.bind("<Button-1>", lambda *_: self._cmd and self._cmd())
        else:
            self.itemconfig(self.rect, fill=DISABLED)
            self.unbind("<Enter>")
            self.unbind("<Leave>")
            self.unbind("<Button-1>")


# ================== Ventana principal ==================


class ExtractorGUI:
    """Ventana principal para extraer datos y exportar a Excel."""

    def __init__(self, root: tk.Tk) -> None:
        if _MISSING_DEPS:
            messagebox.showerror(
                "Dependencias faltantes",
                "No se puede iniciar la GUI porque faltan módulos:\n\n- "
                + "\n- ".join(_MISSING_DEPS)
                + "\n\nRevisa la estructura y vuelve a intentar.",
            )

        self.root = root
        self.root.title(f"{APP_NAME} – {APP_SUBTITLE} {APP_VERSION}")
        self.root.geometry("840x900")
        self.root.resizable(False, False)
        self.root.configure(bg=BG)

        # Estado
        self._filas: list[dict[str, str]] = []
        self._razon_social: str = ""

        # TK vars
        self.var_user = tk.StringVar()
        self.var_pass = tk.StringVar()
        self.var_ot = tk.StringVar()
        self.var_fecha_estimada = tk.StringVar()
        self.var_headless = tk.BooleanVar(value=True)
        self._dev_mode = False

        # Widgets
        self.lbl_prog: tk.Label
        self.progress: ttk.Progressbar
        self.txt_log: scrolledtext.ScrolledText
        self.btn_dev_toggle: ttk.Button
        self.dev_section: tk.Frame | None = None
        self.dev_card: tk.Frame | None = None
        self.btn_dev_bump: ttk.Button | None = None
        self.btn_dev_release: ttk.Button | None = None
        self._dev_card_pack_opts: dict[str, Any] = {}

        self._build()

    # ---------- UI construction ----------
    def _build(self) -> None:
        self._build_header()

        container = tk.Frame(self.root, bg=BG)
        container.pack(fill="both", expand=True, padx=18, pady=12)

        # Herramientas para desarrollador (seccion propia para que aparezca en el lugar correcto)
        self.dev_section = tk.Frame(container, bg=BG)
        self.dev_section.pack(fill="x", pady=(0, 8))
        dev_toggle = tk.Frame(self.dev_section, bg=BG)
        dev_toggle.pack(fill="x")
        self.btn_dev_toggle = ttk.Button(
            dev_toggle,
            text="Modo desarrollador",
            command=self._toggle_dev_mode,
        )
        self.btn_dev_toggle.pack(anchor="w", padx=14)
        self._build_dev_card(self.dev_section)

        # Credenciales
        card_cred = self._card(container, "🔐 Credenciales de acceso")
        self._entry(card_cred, "Usuario MetroWeb", self.var_user)
        self._entry(card_cred, "Contraseña", self.var_pass, show="*")

        # OT
        card_ot = self._card(container, "📋 Orden de Trabajo")
        self._entry(card_ot, "Número de OT (ej. 307-62136)", self.var_ot)
        self._build_fecha_estimada_input(card_ot)
        chk = tk.Checkbutton(
            card_ot,
            text="Ejecutar en modo oculto (headless)",
            variable=self.var_headless,
            bg=CARD,
            activebackground=CARD,
        )
        chk.pack(anchor="w", padx=14, pady=(2, 10))

        # Acciones
        btn_zone = tk.Frame(container, bg=BG)
        btn_zone.pack(pady=6)
        self.btn_run = ModernButton(
            btn_zone, "🚀 INICIAR EXTRACCIÓN", command=self._start_thread
        )
        self.btn_run.pack()

        # Progreso + log
        card_prog = self._card(container, "📊 Progreso y registro")

        top_prog = tk.Frame(card_prog, bg=CARD)
        top_prog.pack(fill="x", padx=14, pady=(10, 6))

        self.progress = ttk.Progressbar(
            top_prog, mode="determinate", length=740, maximum=100
        )
        self.progress.pack(fill="x")
        style = ttk.Style()
        style.theme_use("default")
        style.configure("TProgressbar", troughcolor="#E6EAF2", background=PRIMARY)

        info_prog = tk.Frame(card_prog, bg=CARD)
        info_prog.pack(fill="x", padx=14, pady=(0, 8))
        self.lbl_prog = tk.Label(
            info_prog, text="Listo para iniciar… (0%)", bg=CARD, fg=MUTED
        )
        self.lbl_prog.pack(side="left")

        # Área de log con altura fija
        log_wrap = tk.Frame(card_prog, bg=CARD, height=320)
        log_wrap.pack(fill="both", expand=True, padx=14, pady=(8, 18))
        log_wrap.pack_propagate(False)

        self.txt_log = scrolledtext.ScrolledText(
            log_wrap,
            height=12,
            font=("Consolas", 9),
            bg="#1e1e1e",
            fg="#d4d4d4",
            relief="flat",
        )
        self.txt_log.pack(fill="both", expand=True)
        self._log("Listo.")

        # Card para anexar al Excel base
        card_merge = self._card(
            container, "📎 Anexar la hoja 'datos vpe' a un Excel base"
        )
        tk.Label(
            card_merge,
            text="Agrega la hoja 'datos vpe' como PRIMERA hoja en una COPIA del libro base seleccionado.",
            bg=CARD,
            fg="#444",
        ).pack(anchor="w", padx=14, pady=(10, 6))

        btn_merge = ttk.Button(
            card_merge,
            text="Agregar a Excel base…",
            command=self._cmd_agregar_a_excel_base,
        )
        btn_merge.pack(anchor="w", padx=14, pady=(0, 14))

        # Footer
        tk.Label(
            self.root,
            text="INTI – Instituto Nacional de Tecnología Industrial",
            bg=BG,
            fg="#777",
            font=("Segoe UI", 8),
        ).pack(pady=(0, 8))

    def _build_header(self) -> None:
        head = tk.Frame(self.root, bg=PRIMARY, height=92)
        head.pack(fill="x")
        head.pack_propagate(False)

        left = tk.Frame(head, bg=PRIMARY)
        left.pack(side="left", padx=18, pady=10)

        # Imagen balanza (assets/balanza.png)
        self._img_ref = None
        img_path = Path("assets/balanza.png")

        MAX_W, MAX_H = 64, 64

        def _subsample_factor(w: int, h: int, max_w: int, max_h: int) -> int:
            from math import ceil

            scale = max(w / max_w, h / max_h, 1.0)
            return int(ceil(scale))

        if img_path.exists():
            try:
                from PIL import Image, ImageTk  # type: ignore

                im = Image.open(img_path)
                im.thumbnail((MAX_W, MAX_H), Image.Resampling.LANCZOS)
                self._img_ref = ImageTk.PhotoImage(im)
                tk.Label(left, image=self._img_ref, bg=PRIMARY).pack(
                    side="left", padx=(0, 10)
                )
            except Exception:
                try:
                    raw = tk.PhotoImage(file=str(img_path))
                    factor = _subsample_factor(raw.width(), raw.height(), MAX_W, MAX_H)
                    if factor > 1:
                        raw = raw.subsample(factor, factor)
                    self._img_ref = raw
                    tk.Label(left, image=self._img_ref, bg=PRIMARY).pack(
                        side="left", padx=(0, 10)
                    )
                except Exception:
                    tk.Label(
                        left,
                        text="⚖️",
                        bg=PRIMARY,
                        fg="white",
                        font=("Segoe UI Emoji", 28),
                    ).pack(side="left", padx=(0, 10))
        else:
            tk.Label(
                left, text="⚖️", bg=PRIMARY, fg="white", font=("Segoe UI Emoji", 28)
            ).pack(side="left", padx=(0, 10))

        title_wrap = tk.Frame(left, bg=PRIMARY)
        title_wrap.pack(side="left")
        tk.Label(
            title_wrap,
            text=APP_NAME,
            bg=PRIMARY,
            fg="white",
            font=("Segoe UI", 20, "bold"),
        ).pack(anchor="w", pady=(2, 0))
        tk.Label(
            title_wrap,
            text=APP_SUBTITLE + " – Balanzas para camiones",
            bg=PRIMARY,
            fg="white",
            font=("Segoe UI", 10),
        ).pack(anchor="w")

        # Versión a la derecha
        right = tk.Frame(head, bg=PRIMARY)
        right.pack(side="right", padx=18, pady=10)
        tk.Label(
            right,
            text=f"{APP_VERSION}",
            bg=PRIMARY_DARK,
            fg="white",
            font=("Segoe UI", 10, "bold"),
            padx=10,
            pady=4,
        ).pack()

    def _build_dev_card(self, parent: tk.Widget) -> None:
        self.dev_card = self._card(parent, "Herramientas para desarrollador")
        tk.Label(
            self.dev_card,
            text="Opciones avanzadas para crear releases. Usalas solo si sabes lo que haces.",
            bg=CARD,
            fg="#444",
        ).pack(anchor="w", padx=14, pady=(10, 6))

        btn_wrap = tk.Frame(self.dev_card, bg=CARD)
        btn_wrap.pack(fill="x", pady=(0, 6))

        self.btn_dev_bump = ttk.Button(
            btn_wrap,
            text="Subir version (bump patch)",
            command=self._start_bump_version,
            state=tk.DISABLED,
        )
        self.btn_dev_bump.pack(anchor="w", padx=14, pady=(0, 6))

        self.btn_dev_release = ttk.Button(
            btn_wrap,
            text="Generar copia / release ZIP",
            command=self._start_make_release,
            state=tk.DISABLED,
        )
        self.btn_dev_release.pack(anchor="w", padx=14, pady=(0, 10))

        self._dev_card_pack_opts = {"fill": "x", "pady": (6, 8)}
        if self.dev_card:
            self.dev_card.pack_forget()

    def _card(self, parent: tk.Widget, title: str) -> tk.Frame:
        card = tk.Frame(
            parent, bg=CARD, highlightbackground="#ddd", highlightthickness=1
        )
        card.pack(fill="x", pady=(0, 12))
        tk.Label(
            card, text=title, bg=CARD, fg="#333", font=("Segoe UI", 11, "bold")
        ).pack(anchor="w", padx=14, pady=(14, 8))
        tk.Frame(card, height=1, bg=SEPARATOR).pack(fill="x", padx=14)
        return card

    def _entry(
        self, parent: tk.Widget, label: str, var: tk.StringVar, show: str = ""
    ) -> None:
        wrap = tk.Frame(parent, bg=CARD)
        wrap.pack(fill="x", padx=14, pady=8)
        tk.Label(wrap, text=label + ":", bg=CARD, fg="#555").pack(
            anchor="w", pady=(0, 4)
        )
        border = tk.Frame(
            wrap, bg="#ccc", highlightbackground="#ccc", highlightthickness=1
        )
        border.pack(fill="x")
        e = tk.Entry(
            border,
            textvariable=var,
            bg="white",
            relief="flat",
            font=("Segoe UI", 10),
            show=show,
        )
        e.pack(fill="x", padx=8, pady=8)

    def _build_fecha_estimada_input(self, parent: tk.Widget) -> None:
        wrap = tk.Frame(parent, bg=CARD)
        wrap.pack(fill="x", padx=14, pady=8)
        tk.Label(
            wrap,
            text="Fecha estimada de verificación (dd/mm/aaaa):",
            bg=CARD,
            fg="#555",
        ).pack(anchor="w", pady=(0, 4))

        row = tk.Frame(wrap, bg=CARD)
        row.pack(fill="x")

        border = tk.Frame(
            row, bg="#ccc", highlightbackground="#ccc", highlightthickness=1
        )
        border.pack(side="left", fill="x", expand=True)
        tk.Entry(
            border,
            textvariable=self.var_fecha_estimada,
            bg="white",
            relief="flat",
            font=("Segoe UI", 10),
        ).pack(fill="x", padx=8, pady=8)

        ttk.Button(
            row, text="Pegar fecha", command=self._pegar_fecha_desde_clipboard
        ).pack(side="left", padx=(8, 0))

    # ---------- UX helpers ----------
    def _log(self, msg: str) -> None:
        self.txt_log.config(state="normal")
        ts = datetime.now().strftime("%H:%M:%S")
        self.txt_log.insert("end", f"[{ts}] {msg}\n")
        self.txt_log.see("end")
        self.txt_log.config(state="disabled")
        self.root.update_idletasks()

    def _set_progress_pct(self, pct: float, label: str) -> None:
        pct = max(0.0, min(100.0, pct))
        self.progress["value"] = pct
        self.lbl_prog.config(text=f"{label} ({pct:.0f}%)")
        self.root.update_idletasks()

    def _enable_ui(self, enabled: bool) -> None:
        self.btn_run.set_enabled(enabled)

    def _pegar_fecha_desde_clipboard(self) -> None:
        try:
            contenido = self.root.clipboard_get()
        except Exception:
            messagebox.showwarning(
                "Portapapeles vacío", "No se pudo leer el portapapeles."
            )
            return

        texto = str(contenido).strip()
        if "\n" in texto:
            texto = texto.splitlines()[0].strip()
        if "\t" in texto:
            texto = texto.split("\t", 1)[0].strip()
        self.var_fecha_estimada.set(texto)

    # ---------- Herramientas de desarrollo ----------
    def _toggle_dev_mode(self) -> None:
        self._dev_mode = not self._dev_mode
        if self.dev_card:
            if self._dev_mode:
                self.dev_card.pack(**cast(Mapping[str, Any], self._dev_card_pack_opts))
                self.btn_dev_toggle.config(text="Cerrar modo desarrollador")
            else:
                self.dev_card.pack_forget()
                self.btn_dev_toggle.config(text="Modo desarrollador")
        self._set_dev_actions_enabled(self._dev_mode)

    def _set_dev_actions_enabled(self, enabled: bool) -> None:
        state = tk.NORMAL if enabled else tk.DISABLED
        if self.btn_dev_bump:
            self.btn_dev_bump.config(state=state)
        if self.btn_dev_release:
            self.btn_dev_release.config(state=state)

    def _run_dev_task(self, target: Callable[[], None]) -> None:
        def runner() -> None:
            try:
                self._set_dev_actions_enabled(False)
                target()
            finally:
                self._set_dev_actions_enabled(self._dev_mode)

        threading.Thread(target=runner, daemon=True).start()

    def _leer_version_pyproject(self) -> str | None:
        pyproject = ROOT / "pyproject.toml"
        if not pyproject.exists():
            return None
        m = re.search(
            r'(?m)^\s*version\s*=\s*["\']([^"\']+)["\']',
            pyproject.read_text(encoding="utf-8"),
        )
        if m:
            return m.group(1)
        return None

    def _start_bump_version(self) -> None:
        self._run_dev_task(self._bump_version)

    def _bump_version(self) -> None:
        before = self._leer_version_pyproject()
        self._log("[DEV] Ejecutando tools.bump_version...")
        result = subprocess.run(
            [sys.executable, "-m", "tools.bump_version"],
            cwd=ROOT,
            capture_output=True,
            text=True,
        )
        for line in result.stdout.splitlines():
            self._log(f"[bump] {line}")
        for line in result.stderr.splitlines():
            self._log(f"[bump][err] {line}")

        if result.returncode == 0:
            after = self._leer_version_pyproject()
            mensaje = f"Version actualizada: {before or '-'} -> {after or '-'}"
            messagebox.showinfo("Version actualizada", mensaje)
        else:
            messagebox.showerror(
                "Error al actualizar version",
                f"El comando devolvio {result.returncode}. Revisar el log de errores.",
            )

    def _start_make_release(self) -> None:
        self._run_dev_task(self._make_release)

    def _collect_dist_zips(self) -> set[Path]:
        dist = ROOT / "tools" / "dist"
        if not dist.exists():
            return set()
        return set(dist.glob("*.zip"))

    def _make_release(self) -> None:
        prev = self._collect_dist_zips()
        self._log("[DEV] Generando release (copia ZIP)...")
        result = subprocess.run(
            [sys.executable, "-m", "tools.make_release"],
            cwd=ROOT,
            capture_output=True,
            text=True,
        )
        for line in result.stdout.splitlines():
            self._log(f"[release] {line}")
        for line in result.stderr.splitlines():
            self._log(f"[release][err] {line}")

        if result.returncode == 0:
            nuevos = self._collect_dist_zips() - prev
            zip_generado = None
            if nuevos:
                zip_generado = max(nuevos, key=lambda p: p.stat().st_mtime)
            if zip_generado:
                mensaje = "Copia generada en tools/dist:\n" f"{zip_generado.resolve()}"
                self._log(f"Release creado: {zip_generado.resolve()}")
            else:
                mensaje = (
                    "Release finalizado. Revisa tools/dist para ver el ZIP generado."
                )
            messagebox.showinfo("Release generado", mensaje)
        else:
            messagebox.showerror(
                "Error al generar release",
                f"El comando devolvio {result.returncode}. Revisar el log de errores.",
            )

    # ---------- Validaciones ----------
    def _validate(self) -> bool:
        if not self.var_user.get().strip():
            messagebox.showerror(
                "Falta usuario", "Debes ingresar el usuario de MetroWeb."
            )
            return False
        if not self.var_pass.get().strip():
            messagebox.showerror("Falta contraseña", "Debes ingresar la contraseña.")
            return False
        ot = self.var_ot.get().strip()
        if not ot:
            messagebox.showerror(
                "Falta OT", "Debes ingresar el número de OT (ej. 307-62136)."
            )
            return False
        if not validar_formato_ot(ot):
            if not messagebox.askyesno(
                "Formato no estándar",
                f"El formato '{ot}' no es XXX-XXXXX.\n¿Deseas continuar de todas formas?",
            ):
                return False
        fecha_estimada = self.var_fecha_estimada.get().strip()
        if fecha_estimada and not validar_fecha_ddmmaaaa(fecha_estimada):
            messagebox.showerror(
                "Fecha inválida",
                "La fecha estimada debe tener formato dd/mm/aaaa (ej. 16/12/2025).",
            )
            return False
        return True

    # ---------- Flujo principal ----------
    def _start_thread(self) -> None:
        if not self._validate():
            return
        self.txt_log.config(state="normal")
        self.txt_log.delete("1.0", "end")
        self.txt_log.config(state="disabled")
        self._set_progress_pct(0, "Preparando…")
        self._enable_ui(False)
        threading.Thread(target=self._run, daemon=True).start()

    def _run(self) -> None:
        try:
            user = self.var_user.get().strip()
            pwd = self.var_pass.get().strip()
            ot = self.var_ot.get().strip()
            headless = self.var_headless.get()

            self._set_progress_pct(5, "Iniciando…")
            self._log("Conectando a MetroWeb…")

            def progress_wrapper(idx: int, total: int) -> None:
                base = 45.0
                span = 55.0
                pct = base + (idx / max(1, total)) * span
                self._set_progress_pct(pct, f"Extrayendo instrumentos {idx}/{total}")

            self._set_progress_pct(10, "Autenticando…")
            self._filas = extraer_camiones_por_ot(
                ot=ot,
                user=user,
                pwd=pwd,
                mostrar_navegador=not headless,
                log_callback=self._log,
                progress_callback=progress_wrapper,
            )

            if not self._filas:
                self._log("No se encontraron instrumentos en la OT.")
                messagebox.showwarning(
                    "Sin datos", "No se encontraron instrumentos para la OT indicada."
                )
                return

            self._razon_social = (
                self._filas[0].get("Razón social (Propietario)", "")
                if self._filas
                else ""
            )
            self._set_progress_pct(100, "Extracción completa")
            self._log(f"Extracción completa: {len(self._filas)} instrumento(s).")
            self.root.after(200, self._save_dialog)

        except Exception as e:  # pragma: no cover (UI)
            self._log(f"ERROR: {e}")
            messagebox.showerror("Error en la extracción", str(e))
        finally:
            self._enable_ui(True)

    # ---------- Guardado (plantilla con "datos vpe") ----------
    def _save_dialog(self) -> None:
        ot = self.var_ot.get().strip()
        razon = (
            limpiar_nombre_archivo(self._razon_social)
            if self._razon_social
            else "SIN_RAZON"
        )
        sugerido = f"OT_{ot}_{razon}.xlsx"

        path = filedialog.asksaveasfilename(
            title="Guardar Excel",
            defaultextension=".xlsx",
            initialfile=sugerido,
            filetypes=[("Archivos Excel", "*.xlsx"), ("Todos los archivos", "*.*")],
        )
        if not path:
            self._log("Guardado cancelado por el usuario.")
            messagebox.showinfo("Cancelado", "No se guardó el archivo.")
            self._set_progress_pct(0, "Listo para una nueva extracción")
            return

        try:
            self._log("Rellenando plantilla de Excel…")
            ruta = self._exportar_en_plantilla(Path(path))
            size_kb = ruta.stat().st_size / 1024
            self._log(f"Archivo guardado: {ruta.resolve()} ({size_kb:.2f} KB)")

            if messagebox.askyesno(
                "Éxito",
                f"Archivo creado:\n\n{ruta.name}\n"
                f"Instrumentos: {len(self._filas)}\n"
                f"Tamaño: {size_kb:.2f} KB\n\n"
                "¿Abrir la carpeta contenedora?",
            ):
                self._open_folder(ruta.parent)

        except Exception as e:  # pragma: no cover (UI)
            self._log(f"ERROR al guardar: {e}")
            messagebox.showerror("Error al guardar", str(e))
        finally:
            self._set_progress_pct(0, "Listo para una nueva extracción")

    def _exportar_en_plantilla(self, destino: Path) -> Path:
        if not self._filas:
            raise RuntimeError(
                "Todavía no hay datos de extracción. Ejecutá la extracción primero."
            )

        template_path = TEMPLATE_CAMION_PATH
        if not template_path.exists():
            raise FileNotFoundError(
                f"No se encontró la plantilla base en: {template_path.resolve()}"
            )

        wb = load_workbook(template_path)
        if "datos vpe" not in wb.sheetnames:
            raise ValueError("La plantilla no contiene la hoja 'datos vpe'.")

        ws = wb["datos vpe"]
        instrumento = dict(self._filas[0])
        campo_fecha_estimada = "Fecha estimada de verificación"
        instrumento[campo_fecha_estimada] = self.var_fecha_estimada.get().strip()

        def _asegurar_fila_fecha() -> None:
            for row in ws.iter_rows(min_row=1, max_row=ws.max_row, max_col=1):
                if row[0].value and str(row[0].value) == campo_fecha_estimada:
                    return

            style_col1 = None
            style_col2 = None
            if ws.max_row >= 2:
                template_row = list(ws.iter_rows(min_row=2, max_row=2, max_col=2))[0]
                style_col1 = template_row[0]._style
                if len(template_row) > 1:
                    style_col2 = template_row[1]._style

            ws.insert_rows(2)
            cell_campo = ws.cell(row=2, column=1, value=campo_fecha_estimada)
            cell_valor = ws.cell(row=2, column=2, value="")
            if style_col1:
                cell_campo._style = style_col1
            if style_col2:
                cell_valor._style = style_col2

        _asegurar_fila_fecha()

        # Normalizar fechas relevantes al formato castellano esperado
        for campo in DATE_FIELDS:
            if campo in instrumento:
                instrumento[campo] = _fecha_castellano(str(instrumento.get(campo, "")))

        for row in ws.iter_rows(min_row=1, max_row=ws.max_row, max_col=1):
            cell_campo = row[0]
            campo = cell_campo.value
            if campo is None:
                continue
            valor = instrumento.get(str(campo), "")
            valor_str = "" if valor is None else str(valor)
            ws.cell(row=cell_campo.row, column=2, value=valor_str)  # type: ignore[arg-type]

        destino = destino.with_suffix(".xlsx")
        destino.parent.mkdir(parents=True, exist_ok=True)
        wb.save(destino)
        _restaurar_hojas_desde_template(
            template_path=template_path,
            destino=destino,
            hojas_a_preservar=["Informe"],
        )
        return destino

    # ---------- Anexar a Excel base (botón) ----------
    def _cmd_agregar_a_excel_base(self) -> None:
        if append_sheet_as_first is None:
            messagebox.showwarning(
                "Función no disponible",
                "No se pudo cargar 'excel_merge'. Revisá que exista src/ui/excel_merge.py y los __init__.py.",
            )
            return

        base_path = filedialog.askopenfilename(
            title="Elegir Excel base donde agregar 'datos vpe'",
            filetypes=[("Archivos de Excel", "*.xlsx")],
        )
        if not base_path:
            self._log("Anexado cancelado por el usuario.")
            return

        try:
            df = self._obtener_dataframe_para_exportar()
            out_path = append_sheet_as_first(
                df=df,
                base_xlsx_path=base_path,
                copy_suffix="_con_datos_vpe",
                sheet_base_name="datos vpe",
            )
            messagebox.showinfo(
                "OK",
                "Se creó una COPIA del libro base con la hoja 'datos vpe' como primera.\n\n"
                f"Archivo generado:\n{out_path}",
            )
            self._log(f"Anexado a base: {out_path}")

        except PermissionError as e:
            messagebox.showwarning(
                "Archivo en uso",
                f"{e}\n\nCerrá el Excel y volvé a intentar, o elegí otro archivo.",
            )
        except Exception as e:
            messagebox.showerror("Error al anexar", str(e))

    # ---------- Anexar a Excel base (flujo encadenado) ----------
    def _merge_into_base(self, df) -> None:
        if append_sheet_as_first is None:
            messagebox.showwarning(
                "Función no disponible",
                "No se pudo cargar 'excel_merge'. Revisá que exista src/ui/excel_merge.py y los __init__.py.",
            )
            return

        base_path = filedialog.askopenfilename(
            title="Elegir Excel base donde agregar 'datos vpe'",
            filetypes=[("Archivos de Excel", "*.xlsx")],
        )
        if not base_path:
            self._log("Anexado cancelado por el usuario.")
            return

        try:
            out_path = append_sheet_as_first(
                df=df,
                base_xlsx_path=base_path,
                copy_suffix="_con_datos_vpe",
                sheet_base_name="datos vpe",
            )
            messagebox.showinfo(
                "OK",
                "Se creó una COPIA del libro base con la hoja 'datos vpe' como primera.\n\n"
                f"Archivo generado:\n{out_path}",
            )
            self._log(f"Anexado a base: {out_path}")

        except PermissionError as e:
            messagebox.showwarning(
                "Archivo en uso",
                f"{e}\n\nCerrá el Excel y volvé a intentar, o elegí otro archivo.",
            )
        except Exception as e:
            messagebox.showerror("Error al anexar", str(e))

    def _obtener_dataframe_para_exportar(self):
        """
        Devuelve el DF en formato 2 columnas (Campo | Valor) con separadores
        '=== INSTRUMENTO N ==='. El helper normaliza a 3 columnas (agrega 'Instrumento N').
        """
        if not self._filas:
            raise RuntimeError(
                "Todavía no hay datos de extracción. Ejecutá la extracción primero."
            )
        return armar_hoja_verificacion_2columnas(self._filas)

    @staticmethod
    def _open_folder(folder: Path) -> None:  # pragma: no cover (UI)
        if platform.system() == "Windows":
            os.startfile(folder)  # type: ignore[attr-defined]
        elif platform.system() == "Darwin":
            os.system(f'open "{folder}"')
        else:
            os.system(f'xdg-open "{folder}"')


# ---------- Entry-point ----------
def main() -> None:
    """Inicializa la aplicacion Tk y centra la ventana."""
    root = tk.Tk()
    app = ExtractorGUI(root)

    # Centrar
    root.update_idletasks()
    w, h = root.winfo_width(), root.winfo_height()
    x = (root.winfo_screenwidth() // 2) - (w // 2)
    y = (root.winfo_screenheight() // 2) - (h // 2)
    root.geometry(f"{w}x{h}+{x}+{y}")
    root.mainloop()


if __name__ == "__main__":
    main()
