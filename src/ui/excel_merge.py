# -*- coding: utf-8 -*-
from __future__ import annotations

import os
import re
from pathlib import Path
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment
from openpyxl.utils.dataframe import dataframe_to_rows

def _is_file_locked(path: Path) -> bool:
    try:
        os.rename(str(path), str(path))
        return False
    except (PermissionError, OSError):
        return True

def _next_sheet_name(wb, base_name: str = "datos vpe") -> str:
    names = {ws.title for ws in wb.worksheets}
    if base_name not in names:
        return base_name
    i = 2
    while f"{base_name} ({i})" in names:
        i += 1
    return f"{base_name} ({i})"

def _safe_copy_name(base_path: Path, suffix: str = "_con_datos_vpe") -> Path:
    stem, ext = base_path.stem, base_path.suffix or ".xlsx"
    candidate = base_path.with_name(f"{stem}{suffix}{ext}")
    if not candidate.exists():
        return candidate
    i = 2
    while True:
        candidate = base_path.with_name(f"{stem}{suffix} ({i}){ext}")
        if not candidate.exists():
            return candidate
        i += 1

def _ensure_three_columns(df: pd.DataFrame) -> pd.DataFrame:
    cols = [c.strip().lower() for c in df.columns]
    if set(cols) >= {"campo", "valor", "instrumento n"}:
        return df[["Campo", "Valor", "Instrumento N"]]
    if set(cols) >= {"campo", "valor"}:
        out_rows = []
        current_n = 1
        sep_pat = re.compile(r"=+\s*INSTRUMENTO\s+(\d+)\s*=+", re.IGNORECASE)
        for _, row in df.iterrows():
            campo = str(row.get("Campo", ""))
            valor = row.get("Valor", "")
            m = sep_pat.fullmatch(campo.strip())
            if m:
                current_n = int(m.group(1)); continue
            out_rows.append({"Campo": campo, "Valor": valor, "Instrumento N": current_n})
        return pd.DataFrame(out_rows, columns=["Campo", "Valor", "Instrumento N"])
    raise ValueError("Se esperaban columnas 'Campo' y 'Valor' como mínimo.")

def append_sheet_as_first(
    df: pd.DataFrame,
    base_xlsx_path: str | Path,
    copy_suffix: str = "_con_datos_vpe",
    sheet_base_name: str = "datos vpe",
) -> Path:
    base_path = Path(base_xlsx_path).resolve()
    if not base_path.exists():
        raise FileNotFoundError(f"No se encontró el archivo base: {base_path}")
    if _is_file_locked(base_path):
        raise PermissionError("El archivo base parece estar abierto en Excel. Cerralo o elegí otro destino.")

    wb = load_workbook(filename=str(base_path))
    new_sheet_name = _next_sheet_name(wb, base_name=sheet_base_name)
    ws = wb.create_sheet(title=new_sheet_name)
    wb._sheets.insert(0, wb._sheets.pop(wb._sheets.index(ws)))  # mover a primera

    df3 = _ensure_three_columns(df)

    header_font = Font(bold=True)
    align = Alignment(horizontal="left", vertical="center", wrap_text=True)
    headers = ["Campo", "Valor", "Instrumento N"]
    for j, h in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=j, value=h)
        cell.font = header_font; cell.alignment = align
    for i, row in enumerate(dataframe_to_rows(df3, index=False, header=False), start=2):
        for j, val in enumerate(row, start=1):
            c = ws.cell(row=i, column=j, value=val); c.alignment = align

    ws.column_dimensions["A"].width = 40
    ws.column_dimensions["B"].width = 50
    ws.column_dimensions["C"].width = 18

    copy_path = _safe_copy_name(base_path, suffix=copy_suffix)
    wb.save(str(copy_path))
    return copy_path
