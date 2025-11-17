# -*- coding: utf-8 -*-
"""
Rellena la hoja "Identificación del Instrumento" de la plantilla
a partir del Excel canónico generado por 307-extract_camiones.py,
usando un YAML campo→celda.

Uso interactivo (pregunta rutas):
    python fill_identificacion.py

Uso con argumentos:
    python fill_identificacion.py --raw RUTA_RAW.xlsx \
        --template RUTA_PLANTILLA.xlsx \
        --map maps/identificacion.yml \
        --out RUTA_SALIDA.xlsx \
        [--sheet "Identificación del Instrumento"]
"""

from pathlib import Path
import argparse
import sys
import pandas as pd
from openpyxl import load_workbook

try:
    import yaml
except Exception:
    print("Falta PyYAML. Instalá con: pip install pyyaml")
    sys.exit(1)

DEFAULT_SHEET = "Identificación del Instrumento"


# ============================================================
# Helpers
# ============================================================

def prompt_path(msg: str) -> Path:
    """Pide una ruta existente por consola."""
    while True:
        p = input(f"{msg}: ").strip().strip('"').strip("'")
        if p:
            pth = Path(p)
            if pth.exists():
                return pth
        print("  Ruta inválida. Probá de nuevo.")


def prompt_out_path(sugerido: str = "") -> Path:
    """Pide la ruta de salida y maneja sobrescritura."""
    while True:
        p = input(f"Ruta de salida del Excel RESULTADO (XLSX){' ['+sugerido+']' if sugerido else ''}: ").strip()
        if not p and sugerido:
            p = sugerido
        if not p:
            print("  Ingresá un nombre/ruta .xlsx")
            continue
        pth = Path(p.strip('"').strip("'"))
        if pth.suffix.lower() not in (".xlsx", ".xlsm"):
            print("  Debe terminar en .xlsx/.xlsm")
            continue

        if pth.exists():
            r = input(f"⚠ '{pth.name}' ya existe. ¿Sobrescribir? [s/N]: ").strip().lower()
            if r in ("s", "si", "sí"):
                return pth
            # Si no quiere sobrescribir, sugerimos un nombre alternativo
            base = pth.stem
            suf = 1
            while True:
                cand = pth.with_name(f"{base} ({suf}){pth.suffix}")
                r2 = input(f"¿Usar '{cand.name}'? [S/n]: ").strip().lower()
                if r2 in ("", "s", "si", "sí"):
                    return cand
                p = input("Ingresá otro nombre .xlsx: ").strip()
                if p:
                    pth = Path(p)
                    if not pth.exists():
                        break
            continue
        return pth


def ask_yes_no(msg: str, default: bool = True) -> bool:
    """Pregunta Sí/No simple."""
    hint = "[S/n]" if default else "[s/N]"
    while True:
        r = input(f"{msg} {hint}: ").strip().lower()
        if r == "" and default is not None:
            return default
        if r in ("s", "si", "sí"):
            return True
        if r in ("n", "no"):
            return False
        print("  Respuesta inválida. Ingresá 's' o 'n'.")


def load_map(yaml_path: Path) -> dict:
    """Carga YAML y normaliza estructura."""
    with open(yaml_path, "r", encoding="utf-8") as f:
        mapping = yaml.safe_load(f) or {}
    norm = {}
    for k, v in mapping.items():
        if isinstance(v, str):
            norm[k] = {"cell": v, "keep_existing": False}
        elif isinstance(v, dict) and "cell" in v:
            norm[k] = {"cell": str(v["cell"]), "keep_existing": bool(v.get("keep_existing", False))}
        else:
            raise ValueError(f"Entrada de mapeo inválida para '{k}': {v}")
    return norm


def fill_sheet(template_path: Path, out_path: Path, mapping: dict, record: dict, sheet_name: str):
    """Copia valores del dict a las celdas indicadas en el mapeo."""
    wb = load_workbook(template_path)
    if sheet_name not in wb.sheetnames:
        raise KeyError(f"La hoja '{sheet_name}' no existe en la plantilla. Hojas: {wb.sheetnames}")
    ws = wb[sheet_name]

    for field, cfg in mapping.items():
        cell = cfg["cell"]
        keep_existing = cfg["keep_existing"]
        value = record.get(field, "")

        # Si el valor llega vacío → no escribir (conserva lo existente)
        if value is None or str(value).strip() == "":
            continue

        # Si pide no pisar y la celda ya tiene algo → no escribir
        if keep_existing:
            if ws[cell].value is not None and str(ws[cell].value).strip() != "":
                continue

        ws[cell] = value

    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_path)


# ============================================================
# Main
# ============================================================

def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--raw", help="Excel/CSV con datos canónicos (salida del scraper)")
    ap.add_argument("--template", help="Plantilla XLSX a rellenar")
    ap.add_argument("--map", help="YAML de mapeo campo→celda")
    ap.add_argument("--out", help="Ruta de salida XLSX")
    ap.add_argument("--sheet", default=DEFAULT_SHEET, help=f"Nombre de hoja destino (default: '{DEFAULT_SHEET}')")
    args = ap.parse_args()

    # Solicitar rutas si faltan
    raw_path = Path(args.raw) if args.raw else prompt_path("Ruta del Excel/CSV ORIGEN con los datos a transferir")
    template_path = Path(args.template) if args.template else prompt_path("Ruta del Excel PLANTILLA a rellenar (XLSX)")
    map_path = Path(args.map) if args.map else prompt_path("Ruta del YAML de MAPEO campo→celda")

    # Sugerir nombre de salida basado en la OT (si existe)
    sugerido_out = ""
    try:
        df_preview = pd.read_excel(raw_path) if raw_path.suffix.lower() in (".xlsx", ".xlsm") else pd.read_csv(raw_path)
        if "Numero de O.T." in df_preview.columns:
            ot_val = str(df_preview["Numero de O.T."].iloc[0]).strip()
            if ot_val:
                sugerido_out = str(Path.cwd() / f"OT_{ot_val}_final.xlsx")
    except Exception:
        pass

    out_path = Path(args.out) if args.out else prompt_out_path(sugerido_out)

    # Cargar datos
    if raw_path.suffix.lower() in (".xlsx", ".xlsm"):
        df = pd.read_excel(raw_path)
    elif raw_path.suffix.lower() == ".csv":
        df = pd.read_csv(raw_path)
    else:
        print("El archivo ORIGEN debe ser .xlsx/.xlsm o .csv")
        sys.exit(1)

    if df.empty:
        print("El ORIGEN no tiene filas. Nada para transferir.")
        sys.exit(0)

    mapping = load_map(map_path)

    # Si hay varios instrumentos, ofrecer generar múltiples archivos
    if len(df) > 1:
        multi = ask_yes_no(f"Se detectaron {len(df)} instrumentos. ¿Generar un XLSX por instrumento?", default=True)
    else:
        multi = False

    if multi:
        for idx, row in df.iterrows():
            sufijo = f"_{idx+1}"
            serie = str(row.get("N° de serie Receptor", "")).strip()
            if serie and serie.lower() != "nan":
                sufijo = f"_{serie}"
            destino = out_path.with_name(out_path.stem + sufijo + out_path.suffix)
            fill_sheet(template_path, destino, mapping, row.to_dict(), sheet_name=args.sheet)
            print(f"✔ Generado: {destino}")
    else:
        record = df.iloc[0].to_dict()
        fill_sheet(template_path, out_path, mapping, record, sheet_name=args.sheet)
        print(f"✔ Generado: {out_path}")

    print("Hecho.")


if __name__ == "__main__":
    main()
